"""XLSX export implementation using openpyxl.

Features:
- Bold header row with a light-blue background
- Auto-adjusted column widths (min 10, max 50 characters)
- Frozen top row for easy scrolling
"""

import io
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F5496")  # dark blue
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)

_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 50


class XlsxExporter:
    def export(self, headers: List[str], rows: List[List[Any]]) -> bytes:
        """Serialise *headers* and *rows* to an XLSX workbook in memory."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"

        # Write header row
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT

        # Write data rows
        for row in rows:
            normalised = [
                None if cell is None else cell
                for cell in row
            ]
            ws.append(normalised)

        # Auto-adjust column widths
        col_widths: List[int] = [len(str(h)) for h in headers]
        for row in rows:
            for col_idx, cell in enumerate(row):
                if col_idx < len(col_widths):
                    cell_len = len(str(cell)) if cell is not None else 0
                    col_widths[col_idx] = max(col_widths[col_idx], cell_len)

        for col_idx, width in enumerate(col_widths, start=1):
            clamped = max(_MIN_COL_WIDTH, min(_MAX_COL_WIDTH, width + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = clamped

        # Freeze the header row
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
