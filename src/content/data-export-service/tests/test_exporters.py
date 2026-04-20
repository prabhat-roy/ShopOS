"""Unit tests for the individual exporter classes."""

import csv
import io
import json

import pytest
from openpyxl import load_workbook

from exporter.exporters.csv_exporter import CsvExporter
from exporter.exporters.json_exporter import JsonExporter
from exporter.exporters.pdf_exporter import PdfExporter
from exporter.exporters.xlsx_exporter import XlsxExporter
from exporter.service import ExportService
from exporter.models import ExportFormat, ExportRequest


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

HEADERS = ["id", "name", "email", "amount"]
ROWS = [
    [1, "Alice", "alice@example.com", 99.99],
    [2, "Bob", None, 0.00],
    [3, "Charlie", "charlie@example.com", 250.50],
]

csv_exporter = CsvExporter()
json_exporter = JsonExporter()
xlsx_exporter = XlsxExporter()
pdf_exporter = PdfExporter()
export_service = ExportService()


# ---------------------------------------------------------------------------
# CSV tests
# ---------------------------------------------------------------------------

class TestCsvExporter:
    def test_csv_output_is_parseable(self):
        data = csv_exporter.export(HEADERS, ROWS)
        # Strip BOM before parsing
        text = data.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        parsed_rows = list(reader)
        assert parsed_rows[0] == HEADERS
        assert len(parsed_rows) == len(ROWS) + 1  # header + data rows

    def test_csv_has_utf8_bom(self):
        data = csv_exporter.export(HEADERS, ROWS)
        # UTF-8 BOM is the three bytes EF BB BF
        assert data[:3] == b"\xef\xbb\xbf"

    def test_csv_handles_none_values(self):
        data = csv_exporter.export(["col"], [[None], ["value"]])
        text = data.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        # First data row should have an empty string for None
        assert rows[1][0] == ""

    def test_csv_empty_rows(self):
        data = csv_exporter.export(HEADERS, [])
        text = data.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        parsed = list(reader)
        assert len(parsed) == 1  # only header


# ---------------------------------------------------------------------------
# JSON tests
# ---------------------------------------------------------------------------

class TestJsonExporter:
    def test_json_has_correct_keys(self):
        data = json_exporter.export(HEADERS, ROWS)
        records = json.loads(data.decode("utf-8"))
        assert isinstance(records, list)
        assert len(records) == len(ROWS)
        for record in records:
            for h in HEADERS:
                assert h in record

    def test_json_handles_none_values(self):
        data = json_exporter.export(["col"], [[None]])
        records = json.loads(data.decode("utf-8"))
        assert records[0]["col"] is None

    def test_json_empty_rows(self):
        data = json_exporter.export(HEADERS, [])
        records = json.loads(data.decode("utf-8"))
        assert records == []


# ---------------------------------------------------------------------------
# XLSX tests
# ---------------------------------------------------------------------------

class TestXlsxExporter:
    def test_xlsx_has_header_row(self):
        data = xlsx_exporter.export(HEADERS, ROWS)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        header_row = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
        assert header_row == HEADERS

    def test_xlsx_auto_width_applied(self):
        data = xlsx_exporter.export(HEADERS, ROWS)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            assert width >= 10
            assert width <= 52  # max(50) + 2 padding

    def test_xlsx_row_count_matches(self):
        data = xlsx_exporter.export(HEADERS, ROWS)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        # +1 for header row
        assert ws.max_row == len(ROWS) + 1


# ---------------------------------------------------------------------------
# PDF tests
# ---------------------------------------------------------------------------

class TestPdfExporter:
    def test_pdf_is_non_empty_bytes(self):
        data = pdf_exporter.export(HEADERS, ROWS)
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_pdf_starts_with_pdf_magic(self):
        data = pdf_exporter.export(HEADERS, ROWS)
        assert data[:4] == b"%PDF"

    def test_multi_page_pdf(self):
        """A PDF with many rows should produce more bytes than a single-row one."""
        many_rows = [[i, f"name_{i}", f"email_{i}@x.com", i * 1.5] for i in range(500)]
        data_many = pdf_exporter.export(HEADERS, many_rows)
        data_one = pdf_exporter.export(HEADERS, ROWS[:1])
        assert len(data_many) > len(data_one)


# ---------------------------------------------------------------------------
# ExportService tests
# ---------------------------------------------------------------------------

class TestExportService:
    def test_row_count_limit_enforced(self, monkeypatch):
        from exporter import config as cfg
        monkeypatch.setattr(cfg.settings, "MAX_ROWS", 2)

        request = ExportRequest(
            format=ExportFormat.CSV,
            filename="test",
            headers=HEADERS,
            rows=ROWS,  # 3 rows > limit of 2
        )
        with pytest.raises(ValueError, match="exceeds the maximum allowed limit"):
            export_service.export(request)

    def test_empty_rows_export(self):
        request = ExportRequest(
            format=ExportFormat.JSON,
            filename="empty",
            headers=HEADERS,
            rows=[],
        )
        data, content_type, filename = export_service.export(request)
        assert b"[]" in data
        assert filename == "empty.json"
