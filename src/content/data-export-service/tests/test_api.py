"""Integration tests for the Data Export Service FastAPI endpoints."""

import csv
import io
import json

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from main import app

client = TestClient(app)

# ---------------------------------------------------------------------------
# Shared payload helpers
# ---------------------------------------------------------------------------

HEADERS = ["id", "product", "price", "in_stock"]
ROWS = [
    [1, "Widget A", 9.99, True],
    [2, "Widget B", 19.99, False],
    [3, "Gadget X", 49.99, True],
]


def _payload(fmt: str, filename: str = "test_export", extra: dict | None = None) -> dict:
    body = {
        "format": fmt,
        "filename": filename,
        "headers": HEADERS,
        "rows": ROWS,
    }
    if extra:
        body.update(extra)
    return body


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestHealthz:
    def test_healthz_returns_200(self):
        response = client.get("/healthz")
        assert response.status_code == 200
        assert response.json() == {"status": "ok"}


class TestExportEndpoint:
    def test_export_csv_via_generic_endpoint(self):
        response = client.post("/export", json=_payload("CSV"))
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert 'attachment; filename="test_export.csv"' in response.headers["content-disposition"]

    def test_export_json_via_generic_endpoint(self):
        response = client.post("/export", json=_payload("JSON"))
        assert response.status_code == 200
        assert "application/json" in response.headers["content-type"]
        records = json.loads(response.content)
        assert len(records) == len(ROWS)
        assert records[0]["id"] == 1


class TestCsvEndpoint:
    def test_csv_export_returns_200_with_correct_content_type(self):
        response = client.post("/export/csv", json=_payload("CSV"))
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]

    def test_csv_content_is_parseable(self):
        response = client.post("/export/csv", json=_payload("CSV"))
        text = response.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        assert rows[0] == HEADERS
        assert len(rows) == len(ROWS) + 1


class TestJsonEndpoint:
    def test_json_export_returns_200(self):
        response = client.post("/export/json", json=_payload("JSON"))
        assert response.status_code == 200

    def test_json_content_has_correct_structure(self):
        response = client.post("/export/json", json=_payload("JSON"))
        records = json.loads(response.content)
        assert isinstance(records, list)
        assert all(h in records[0] for h in HEADERS)


class TestXlsxEndpoint:
    def test_xlsx_export_returns_200(self):
        response = client.post("/export/xlsx", json=_payload("XLSX"))
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

    def test_xlsx_content_has_correct_header_row(self):
        response = client.post("/export/xlsx", json=_payload("XLSX"))
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        header_row = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
        assert header_row == HEADERS


class TestPdfEndpoint:
    def test_pdf_export_returns_200(self):
        response = client.post("/export/pdf", json=_payload("PDF", extra={"title": "Test Report"}))
        assert response.status_code == 200
        assert "application/pdf" in response.headers["content-type"]

    def test_pdf_content_is_valid_pdf(self):
        response = client.post("/export/pdf", json=_payload("PDF"))
        assert response.content[:4] == b"%PDF"


class TestValidationErrors:
    def test_exceeds_max_rows_returns_400(self, monkeypatch):
        from exporter import config as cfg
        monkeypatch.setattr(cfg.settings, "MAX_ROWS", 1)

        response = client.post("/export", json=_payload("CSV"))
        assert response.status_code == 400
        assert "exceeds" in response.json()["detail"]

    def test_missing_headers_returns_422(self):
        payload = {
            "format": "CSV",
            "filename": "bad",
            "rows": [[1, 2, 3]],
            # "headers" intentionally omitted
        }
        response = client.post("/export", json=payload)
        assert response.status_code == 422
